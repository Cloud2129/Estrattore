"""
Estrattore Pratiche by Jurij
Versione con: tema scuro, splash screen, PDF, avviso scadenze,
aggiornamento automatico, font regolabile, bookmarklet.

Requisiti:
    pip install customtkinter openpyxl reportlab
"""

import customtkinter as ctk
from tkinter import messagebox, filedialog
import tkinter as tk
from tkinter import ttk
import threading
import json
import os
import sys
import base64
from http.server import HTTPServer, BaseHTTPRequestHandler
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
#  CONFIGURAZIONE
# ─────────────────────────────────────────────
APP_NOME    = "Estrattore Pratiche by Jurij"
APP_VERSION = "2.0"

try:
    OPERATORE = os.getlogin()
except Exception:
    OPERATORE = os.environ.get("USERNAME", "OPERATORE")

EXCEL_PATH_DEFAULT = f"pratiche_{OPERATORE}.xlsx"
SERVER_PORT        = 7432
SETTINGS_FILE      = "settings.json"

# ─────────────────────────────────────────────
#  SETTINGS (tema + font)
# ─────────────────────────────────────────────
def carica_settings() -> dict:
    try:
        if os.path.exists(SETTINGS_FILE):
            return json.loads(open(SETTINGS_FILE).read())
    except Exception:
        pass
    return {"tema": "light", "font_size": 12}

def salva_settings(s: dict):
    try:
        open(SETTINGS_FILE, "w").write(json.dumps(s))
    except Exception:
        pass

SETTINGS = carica_settings()

# ─────────────────────────────────────────────
#  COLONNE EXCEL
# ─────────────────────────────────────────────
COL_PRATICHE = [
    "Nr. Pratica", "Operatore Gestionale", "Stato Pratica",
    "Tipo Soggiorno", "Tipo Pratica", "Validita Soggiorno",
    "Cognome", "Nome", "CUI", "Sesso", "Data Nascita",
    "Luogo Nascita", "Nazione Nascita", "Cittadinanza",
    "Codice Fiscale", "Stato Civile", "Telefono",
    "Comune", "Indirizzo", "Motivo Soggiorno",
    "Coniuge", "Referenze", "Note Gestionale",
    "Data Presentazione", "Scadenza Rinnovo",
    "Estratto Da", "Data Estrazione", "Data Appuntamento",
]

COL_ATTIVITA = [
    "Nr. Pratica", "Stato",
    "SDI Esito", "SDI Note",
    "Reddito Tipo", "Reddito Note",
    "Famiglia Trainante", "Nome Trainante", "CF Trainante", "Nr Pratica Trainante",
    "Documenti Mancanti", "Note Personali",
    "Checklist Compilata Da", "Checklist Compilata Il", "Note Modificate Il",
]

STATI = {
    "da_verificare": ("🟡 Da verificare", "#FDE68A", "#B45309"),
    "sospesa":       ("🟠 Sospesa",       "#FCD34D", "#C2410C"),
    "validata":      ("🟢 Validata",      "#86EFAC", "#15803D"),
    "negata":        ("🔴 Negata",        "#FCA5A5", "#B91C1C"),
}

SDI_ESITI    = ["", "Positivo", "Negativo", "Positivo non ostativo"]
REDDITO_TIPI = ["", "Lavoro subordinato", "Lavoro autonomo", "Nessun reddito", "Altro"]

# ─────────────────────────────────────────────
#  TEMA (chiaro / scuro)
# ─────────────────────────────────────────────
TEMI = {
    "light": {
        "blu_scuro":   "#1F4E79",
        "blu_medio":   "#2E75B6",
        "sfondo":      "#F0F4F8",
        "card":        "#FFFFFF",
        "testo":       "#1E293B",
        "testo_light": "#64748B",
        "bordo":       "#E2E8F0",
        "verde":       "#16A34A",
        "verde_dark":  "#15803D",
        "arancio":     "#EA580C",
        "top_bg":      "#1F4E79",
        "top_text":    "white",
        "top_sub":     "#93C5FD",
        "nav_bg":      "#FFFFFF",
        "nav_border":  "#E2E8F0",
        "tree_bg":     "#FFFFFF",
        "tree_fg":     "#1E293B",
        "tree_sel":    "#2E75B6",
    },
    "dark": {
        "blu_scuro":   "#0F2847",
        "blu_medio":   "#1D4ED8",
        "sfondo":      "#1E2433",
        "card":        "#252D3D",
        "testo":       "#E2E8F0",
        "testo_light": "#94A3B8",
        "bordo":       "#334155",
        "verde":       "#16A34A",
        "verde_dark":  "#15803D",
        "arancio":     "#EA580C",
        "top_bg":      "#0F2847",
        "top_text":    "#E2E8F0",
        "top_sub":     "#60A5FA",
        "nav_bg":      "#252D3D",
        "nav_border":  "#334155",
        "tree_bg":     "#252D3D",
        "tree_fg":     "#E2E8F0",
        "tree_sel":    "#1D4ED8",
    }
}

def T() -> dict:
    return TEMI["light"]

def FS(delta=0) -> int:
    return max(9, min(16, SETTINGS.get("font_size", 12) + delta))

# ─────────────────────────────────────────────
#  BOOKMARKLET
# ─────────────────────────────────────────────
BOOKMARKLET_JS = r"""javascript:(function(){{
  function vt(t){{var e=document.querySelector("input[title='"+t+"'],textarea[title='"+t+"']");return e?(e.value||e.innerText||"").trim():""}}
  function vi(id){{var e=document.getElementById(id);return e?(e.value||e.innerText||"").trim():""}}
  function vl(lbl){{
    var l=Array.from(document.querySelectorAll("label")).find(function(x){{return x.innerText.trim().replace(":","").trim().toLowerCase()===lbl.toLowerCase()}});
    if(!l)return"";var td=l.closest("td");if(!td)return"";var n=td.nextElementSibling;if(!n)return"";var i=n.querySelector("input");return i?(i.value||"").trim():""
  }}
  var h3=document.querySelector("h3"),h3t=h3?h3.innerText:"";
  var m1=h3t.match(/Pratica[\s]+n[\xb0o][\s]*(\S+)/i),m2=h3t.match(/assegnata all'utente[\s]+(\S+)/i);
  var ne=document.getElementById("idnotexx");
  var d={{"Nr. Pratica":m1?m1[1]:"","Operatore Gestionale":m2?m2[1]:"","Stato Pratica":vi("statoPraticaDescrConv"),"Tipo Soggiorno":vi("docSoggiorno"),"Tipo Pratica":vl("Tipo Pratica"),"Validita Soggiorno":vt("Validit\u00e0 del soggiorno"),"Cognome":vl("Cognome"),"Nome":vl("Nome"),"CUI":vt("cui"),"Sesso":vl("Sesso"),"Data Nascita":vi("dataNascitaStraniero"),"Luogo Nascita":vt("Luogo di Nascita"),"Nazione Nascita":vl("Nazione Nascita"),"Cittadinanza":vt("Cittadinanza dello straniero"),"Codice Fiscale":vt("Codice fiscale dello straniero"),"Stato Civile":vt("Stato Civile"),"Telefono":vi("telefono"),"Comune":vt("Comune di residenza in Italia"),"Indirizzo":vt("Indirizzo"),"Motivo Soggiorno":vi("motivoSoggiorno"),"Coniuge":vt("Cognome e\/o Nome del coniuge"),"Referenze":vt("Eventuali referenze"),"Note Gestionale":ne?(ne.value||ne.innerText||"").trim():"","Data Presentazione":vt("data di presentazione della istanza"),"Scadenza Rinnovo":vt("Data scadenza rinnovo")}};
  if(!d["Nr. Pratica"]){{alert("Nessuna pratica trovata.");return}}
  fetch("http://localhost:{port}/pratica",{{method:"POST",headers:{{"Content-Type":"application/json"}},body:JSON.stringify(d)}}).then(function(r){{if(r.ok)alert("\u2705 "+d["Nr. Pratica"]+" inviata!");else alert("\u274c Errore. App aperta?");}}).catch(function(){{alert("\u274c Impossibile connettersi. Pratiche.exe aperto?");}});
}})();""".format(port=SERVER_PORT)

# ─────────────────────────────────────────────
#  SERVER HTTP LOCALE
# ─────────────────────────────────────────────
class _Handler(BaseHTTPRequestHandler):
    callback = None
    def do_POST(self):
        if self.path == "/pratica":
            try:
                body = self.rfile.read(int(self.headers.get("Content-Length",0)))
                dati = json.loads(body.decode("utf-8"))
                self.send_response(200)
                self.send_header("Content-Type","application/json")
                self.send_header("Access-Control-Allow-Origin","*")
                self.end_headers()
                self.wfile.write(b'{"ok":true}')
                if _Handler.callback:
                    _Handler.callback(dati)
            except Exception:
                self.send_response(500); self.end_headers()
        else:
            self.send_response(404); self.end_headers()
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin","*")
        self.send_header("Access-Control-Allow-Methods","POST,OPTIONS")
        self.send_header("Access-Control-Allow-Headers","Content-Type")
        self.end_headers()
    def log_message(self,*a): pass

class ServerLocale:
    def __init__(self, callback):
        _Handler.callback = callback
        self._server = HTTPServer(("localhost", SERVER_PORT), _Handler)
        self._thread = threading.Thread(target=self._server.serve_forever, daemon=True)
    def avvia(self): self._thread.start()
    def ferma(self): self._server.shutdown()

# ─────────────────────────────────────────────
#  LOGICA EXCEL
# ─────────────────────────────────────────────
def _bordo():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _intestazione(ws, colonne, colore="1F4E79"):
    for col, nome in enumerate(colonne, start=1):
        c = ws.cell(row=1, column=col, value=nome)
        c.font      = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
        c.fill      = PatternFill("solid", start_color=colore)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _bordo()
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

def crea_o_apri_excel(path):
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    ws_p = wb.active; ws_p.title = "Pratiche"
    _intestazione(ws_p, COL_PRATICHE, "1F4E79")
    for i,w in enumerate([14,20,14,28,18,14,16,16,12,6,12,14,16,16,18,14,14,16,28,22,20,22,30,14,14,16,18],1):
        ws_p.column_dimensions[get_column_letter(i)].width = w
    ws_a = wb.create_sheet("Attivita")
    _intestazione(ws_a, COL_ATTIVITA, "1E3A5F")
    for i,w in enumerate([14,18,20,30,20,30,16,22,18,16,35,40,20,18,18],1):
        ws_a.column_dimensions[get_column_letter(i)].width = w
    wb.save(path); return wb

def _trova_riga(ws, nr, col=1):
    for row in ws.iter_rows(min_row=2):
        if str(row[col-1].value or "").strip() == nr.strip():
            return row[0].row
    return None

def _scrivi_riga(ws, riga, colonne, dati):
    fill = PatternFill("solid", start_color="EFF6FF") if riga%2==0 else None
    for col, nome in enumerate(colonne, start=1):
        c = ws.cell(row=riga, column=col, value=dati.get(nome,""))
        c.font = Font(name="Segoe UI", size=10)
        c.alignment = Alignment(vertical="center")
        c.border = _bordo()
        if fill: c.fill = fill

def salva_pratica_excel(path, dati):
    wb = crea_o_apri_excel(path)
    ws = wb["Pratiche"]
    nr = dati.get("Nr. Pratica","").strip()
    riga = _trova_riga(ws, nr) or (ws.max_row+1)
    _scrivi_riga(ws, riga, COL_PRATICHE, dati)
    ws_a = wb["Attivita"]
    if not _trova_riga(ws_a, nr):
        _scrivi_riga(ws_a, ws_a.max_row+1, COL_ATTIVITA, {"Nr. Pratica":nr,"Stato":"da_verificare"})
    wb.save(path)

def leggi_pratiche(path):
    if not os.path.exists(path): return []
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Pratiche" not in wb.sheetnames: return []
    ws = wb["Pratiche"]; hdr = [c.value for c in ws[1]]
    return [dict(zip(hdr,r)) for r in ws.iter_rows(min_row=2,values_only=True) if any(r)]

def leggi_attivita(path, nr):
    if not os.path.exists(path): return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Attivita" not in wb.sheetnames: return {}
    ws = wb["Attivita"]; hdr = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr,row))
        if str(d.get("Nr. Pratica","")).strip() == nr.strip(): return d
    return {}

def leggi_tutti_stati(path):
    if not os.path.exists(path): return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Attivita" not in wb.sheetnames: return {}
    ws = wb["Attivita"]; hdr = [c.value for c in ws[1]]; out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr,row)); nr = str(d.get("Nr. Pratica","")).strip()
        if nr: out[nr] = str(d.get("Stato","da_verificare") or "da_verificare")
    return out

def salva_attivita_excel(path, dati):
    wb = openpyxl.load_workbook(path)
    ws = wb["Attivita"]; hdr = [c.value for c in ws[1]]
    nr = str(dati.get("Nr. Pratica","")).strip()
    riga = _trova_riga(ws,nr) or (ws.max_row+1)
    fill = PatternFill("solid", start_color="EFF6FF") if riga%2==0 else None
    for col,nome in enumerate(hdr,start=1):
        c = ws.cell(row=riga,column=col,value=dati.get(nome,""))
        c.font=Font(name="Segoe UI",size=10); c.alignment=Alignment(vertical="center"); c.border=_bordo()
        if fill: c.fill=fill
    wb.save(path)

def pratiche_in_scadenza(path, giorni=30):
    pratiche = leggi_pratiche(path)
    oggi = date.today(); limite = oggi + timedelta(days=giorni)
    out = []
    for p in pratiche:
        try:
            scad = datetime.strptime(str(p.get("Scadenza Rinnovo","")),"%d/%m/%Y").date()
            if oggi <= scad <= limite: out.append(p)
        except Exception: pass
    return out

# ─────────────────────────────────────────────
#  GENERAZIONE PDF
# ─────────────────────────────────────────────
def genera_pdf(pratica: dict, att: dict, path_out: str):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)

    doc  = SimpleDocTemplate(path_out, pagesize=A4,
                              leftMargin=2*cm, rightMargin=2*cm,
                              topMargin=2*cm, bottomMargin=2*cm)
    stili = getSampleStyleSheet()

    BLU   = colors.HexColor("#1F4E79")
    BLU2  = colors.HexColor("#2E75B6")
    GRIG  = colors.HexColor("#F0F4F8")
    BORD  = colors.HexColor("#E2E8F0")

    s_titolo = ParagraphStyle("titolo", fontSize=16, textColor=BLU,
                               fontName="Helvetica-Bold", spaceAfter=4)
    s_sub    = ParagraphStyle("sub",    fontSize=10, textColor=colors.HexColor("#64748B"),
                               fontName="Helvetica", spaceAfter=12)
    s_sez    = ParagraphStyle("sez",    fontSize=11, textColor=BLU2,
                               fontName="Helvetica-Bold", spaceBefore=10, spaceAfter=4)
    s_lbl    = ParagraphStyle("lbl",    fontSize=8,  textColor=colors.HexColor("#64748B"),
                               fontName="Helvetica")
    s_val    = ParagraphStyle("val",    fontSize=10, textColor=colors.HexColor("#1E293B"),
                               fontName="Helvetica")
    s_note   = ParagraphStyle("note",   fontSize=10, textColor=colors.HexColor("#1E293B"),
                               fontName="Helvetica", leading=14)

    storia = []

    # Intestazione
    nr   = str(pratica.get("Nr. Pratica","") or "")
    nome = f"{pratica.get('Cognome','')} {pratica.get('Nome','')}".strip()
    storia.append(Paragraph(f"Scheda Pratica — {nr}", s_titolo))
    storia.append(Paragraph(
        f"{nome}  •  Estratto da {pratica.get('Estratto Da','')}  •  {pratica.get('Data Estrazione','')}",
        s_sub))
    storia.append(HRFlowable(width="100%", thickness=2, color=BLU))
    storia.append(Spacer(1, 0.3*cm))

    # Stato
    stato_key = str(att.get("Stato","da_verificare") or "da_verificare")
    info_stato = STATI.get(stato_key, STATI["da_verificare"])
    stato_colore = colors.HexColor(info_stato[2])
    storia.append(Paragraph(f"Stato: <font color='#{info_stato[2][1:]}'><b>{info_stato[0]}</b></font>", s_val))
    storia.append(Spacer(1, 0.3*cm))

    def tabella_campi(campi):
        righe = []
        riga_corrente = []
        for i, (lbl, val) in enumerate(campi):
            cella = [Paragraph(lbl, s_lbl), Paragraph(str(val or "—"), s_val)]
            riga_corrente.append(cella)
            if len(riga_corrente) == 2:
                righe.append(riga_corrente)
                riga_corrente = []
        if riga_corrente:
            riga_corrente.append(["",""])
            righe.append(riga_corrente)

        # Flatten per Table
        dati_tab = []
        for riga in righe:
            r = []
            for cella in riga:
                r.extend(cella)
            dati_tab.append(r)

        t = Table(dati_tab, colWidths=[3*cm, 6.5*cm, 3*cm, 6.5*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), GRIG),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, GRIG]),
            ("GRID", (0,0), (-1,-1), 0.5, BORD),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
        ]))
        return t

    # Anagrafica
    storia.append(Paragraph("Anagrafica", s_sez))
    storia.append(tabella_campi([
        ("Cognome",         pratica.get("Cognome")),
        ("Nome",            pratica.get("Nome")),
        ("CUI",             pratica.get("CUI")),
        ("Codice Fiscale",  pratica.get("Codice Fiscale")),
        ("Data Nascita",    pratica.get("Data Nascita")),
        ("Luogo Nascita",   pratica.get("Luogo Nascita")),
        ("Nazione Nascita", pratica.get("Nazione Nascita")),
        ("Cittadinanza",    pratica.get("Cittadinanza")),
        ("Sesso",           pratica.get("Sesso")),
        ("Stato Civile",    pratica.get("Stato Civile")),
        ("Telefono",        pratica.get("Telefono")),
        ("Coniuge",         pratica.get("Coniuge")),
    ]))

    storia.append(Spacer(1, 0.3*cm))
    storia.append(Paragraph("Soggiorno e Residenza", s_sez))
    storia.append(tabella_campi([
        ("Tipo Soggiorno",   pratica.get("Tipo Soggiorno")),
        ("Tipo Pratica",     pratica.get("Tipo Pratica")),
        ("Motivo Soggiorno", pratica.get("Motivo Soggiorno")),
        ("Validita Sogg.",   pratica.get("Validita Soggiorno")),
        ("Presentazione",    pratica.get("Data Presentazione")),
        ("Scadenza Rinnovo", pratica.get("Scadenza Rinnovo")),
        ("Comune",           pratica.get("Comune")),
        ("Indirizzo",        pratica.get("Indirizzo")),
        ("Referenze",        pratica.get("Referenze")),
        ("Op. Gestionale",   pratica.get("Operatore Gestionale")),
    ]))

    # Checklist
    storia.append(Spacer(1, 0.3*cm))
    storia.append(Paragraph("Checklist", s_sez))
    storia.append(tabella_campi([
        ("SDI Esito",           att.get("SDI Esito")),
        ("Reddito Tipo",        att.get("Reddito Tipo")),
        ("Famiglia Trainante",  att.get("Famiglia Trainante")),
        ("Nome Trainante",      att.get("Nome Trainante")),
        ("CF Trainante",        att.get("CF Trainante")),
        ("Nr Pratica Trainante",att.get("Nr Pratica Trainante")),
    ]))

    if att.get("SDI Note"):
        storia.append(Spacer(1,0.2*cm))
        storia.append(Paragraph(f"<b>Note SDI:</b> {att.get('SDI Note','')}", s_note))
    if att.get("Reddito Note"):
        storia.append(Paragraph(f"<b>Note Reddito:</b> {att.get('Reddito Note','')}", s_note))
    if att.get("Documenti Mancanti"):
        storia.append(Paragraph(f"<b>Documenti mancanti:</b> {att.get('Documenti Mancanti','')}", s_note))

    # Note personali
    if att.get("Note Personali"):
        storia.append(Spacer(1,0.3*cm))
        storia.append(Paragraph("Note Personali", s_sez))
        storia.append(Paragraph(str(att.get("Note Personali","")), s_note))

    # Note gestionale
    if pratica.get("Note Gestionale"):
        storia.append(Spacer(1,0.3*cm))
        storia.append(Paragraph("Note Gestionale", s_sez))
        storia.append(Paragraph(str(pratica.get("Note Gestionale","")), s_note))

    # Footer
    storia.append(Spacer(1,0.5*cm))
    storia.append(HRFlowable(width="100%", thickness=1, color=BORD))
    storia.append(Paragraph(
        f"Estrattore Pratiche by Jurij  •  Stampato il {datetime.now().strftime('%d/%m/%Y %H:%M')}  •  {OPERATORE}",
        ParagraphStyle("footer", fontSize=8, textColor=colors.HexColor("#94A3B8"),
                       fontName="Helvetica", alignment=1)))
    doc.build(storia)

# ─────────────────────────────────────────────
#  SPLASH SCREEN
# ─────────────────────────────────────────────
class SplashScreen(tk.Toplevel):
    def __init__(self, root):
        super().__init__(root)
        self.overrideredirect(True)
        w, h = 420, 220
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self.configure(bg="#1F4E79")

        tk.Label(self, text="🏛", font=("Segoe UI", 36),
                 bg="#1F4E79", fg="white").pack(pady=(30,4))
        tk.Label(self, text=APP_NOME,
                 font=("Segoe UI", 14, "bold"),
                 bg="#1F4E79", fg="white").pack()
        tk.Label(self, text=f"v{APP_VERSION}  —  {OPERATORE}",
                 font=("Segoe UI", 10),
                 bg="#1F4E79", fg="#93C5FD").pack(pady=4)

        self._bar_frame = tk.Frame(self, bg="#1F4E79")
        self._bar_frame.pack(pady=16)
        self._bar_bg = tk.Frame(self._bar_frame, bg="#0F2847", width=300, height=6)
        self._bar_bg.pack()
        self._bar_fg = tk.Frame(self._bar_frame, bg="#60A5FA", width=0, height=6)
        self._bar_fg.place(x=0, y=0)
        self._progress = 0
        self._anima()

    def _anima(self):
        self._progress += 4
        self._bar_fg.configure(width=min(300, int(300 * self._progress / 100)))
        if self._progress < 100:
            self.after(30, self._anima)

    def chiudi(self):
        self.destroy()

# ─────────────────────────────────────────────
#  BANNER SCADENZE
# ─────────────────────────────────────────────
class BannerScadenze(ctk.CTkFrame):
    def __init__(self, parent, pratiche_scad, on_click, **kw):
        T_ = T()
        super().__init__(parent, fg_color="#FEF3C7",
                         corner_radius=0, height=36, **kw)
        self.pack_propagate(False)
        n = len(pratiche_scad)
        ctk.CTkLabel(self, text=f"  ⚠️  {n} pratica{'e' if n>1 else ''} in scadenza entro 30 giorni",
                     font=ctk.CTkFont(size=11, weight="bold"),
                     text_color="#B45309",
                     fg_color="transparent").pack(side="left", padx=8)
        ctk.CTkButton(self, text="Vedi →",
                      command=on_click,
                      fg_color="#F59E0B", hover_color="#D97706",
                      text_color="white", height=24, width=70,
                      corner_radius=6, font=ctk.CTkFont(size=11)
                      ).pack(side="left", padx=4)
        ctk.CTkButton(self, text="✕",
                      command=self.destroy,
                      fg_color="transparent", text_color="#B45309",
                      hover_color="#FDE68A", width=28, height=24,
                      corner_radius=6, font=ctk.CTkFont(size=12)
                      ).pack(side="right", padx=4)

# ─────────────────────────────────────────────
#  POPUP ANTEPRIMA
# ─────────────────────────────────────────────
class PopupAnteprima(ctk.CTkToplevel):
    def __init__(self, parent, dati, on_conferma):
        super().__init__(parent)
        T_ = T()
        self.title("Anteprima estrazione")
        self.geometry("460x380")
        self.resizable(False, False)
        self.grab_set(); self.lift(); self.focus_force()
        self._dati = dati; self._on_conferma = on_conferma
        self.configure(fg_color=T_["sfondo"])
        self._build(dati)

    def _build(self, d):
        T_ = T()
        hdr = ctk.CTkFrame(self, fg_color=T_["blu_scuro"], corner_radius=0, height=52)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        nome = f"{d.get('Cognome','')} {d.get('Nome','')}".strip() or "—"
        ctk.CTkLabel(hdr, text=f"  📋  {nome}",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=T_["top_text"]).pack(side="left", padx=14)
        ctk.CTkLabel(hdr, text=d.get("Nr. Pratica",""),
                     font=ctk.CTkFont(family="Consolas", size=12),
                     text_color=T_["top_sub"]).pack(side="right", padx=14)

        card = ctk.CTkFrame(self, fg_color=T_["card"], corner_radius=10,
                            border_color=T_["bordo"], border_width=1)
        card.pack(fill="both", expand=True, padx=14, pady=12)
        g = ctk.CTkFrame(card, fg_color="transparent")
        g.pack(fill="both", expand=True, padx=14, pady=10)
        g.columnconfigure(1, weight=1); g.columnconfigure(3, weight=1)
        for i,(lbl,val) in enumerate([
            ("CUI",d.get("CUI")),("Codice Fiscale",d.get("Codice Fiscale")),
            ("Nazione",d.get("Nazione Nascita")),("Data Nascita",d.get("Data Nascita")),
            ("Tipo Soggiorno",d.get("Tipo Soggiorno")),("Motivo",d.get("Motivo Soggiorno")),
            ("Comune",d.get("Comune")),("Scadenza",d.get("Scadenza Rinnovo")),
        ]):
            r,cl = i//2,(i%2)*2
            ctk.CTkLabel(g, text=lbl, font=ctk.CTkFont(size=10),
                         text_color=T_["testo_light"]).grid(row=r,column=cl,sticky="w",padx=(0,6),pady=3)
            ctk.CTkLabel(g, text=str(val or "—"),
                         font=ctk.CTkFont(size=FS(), weight="bold"),
                         text_color=T_["testo"]).grid(row=r,column=cl+1,sticky="w",pady=3,padx=(0,16))

        br = ctk.CTkFrame(self, fg_color=T_["sfondo"], corner_radius=0)
        br.pack(fill="x", padx=14, pady=(0,14))
        ctk.CTkButton(br, text="✖  Annulla", command=self.destroy,
                      fg_color=T_["bordo"], text_color=T_["testo"],
                      hover_color=T_["sfondo"], corner_radius=8, height=36,
                      font=ctk.CTkFont(size=12)).pack(side="left", padx=(0,8))
        ctk.CTkButton(br, text="✅  Conferma e Salva", command=self._conferma,
                      fg_color=T_["verde"], hover_color=T_["verde_dark"],
                      corner_radius=8, height=36,
                      font=ctk.CTkFont(size=12, weight="bold")
                      ).pack(side="left", fill="x", expand=True)

    def _conferma(self):
        self.destroy(); self._on_conferma(self._dati)

# ─────────────────────────────────────────────
#  POPUP BOOKMARKLET
# ─────────────────────────────────────────────
class PopupBookmarklet(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        T_ = T()
        self.title("Bookmarklet")
        self.geometry("600x440")
        self.resizable(False, False)
        self.lift(); self.focus_force()
        self.configure(fg_color=T_["sfondo"])
        self._build()

    def _build(self):
        T_ = T()
        hdr = ctk.CTkFrame(self, fg_color=T_["blu_scuro"], corner_radius=0, height=52)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="  📎  Configura il Bookmarklet",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=T_["top_text"]).pack(side="left", padx=14)

        body = ctk.CTkFrame(self, fg_color=T_["sfondo"], corner_radius=0)
        body.pack(fill="both", expand=True, padx=16, pady=14)

        ctk.CTkLabel(body,
                     text="Aggiungi il bookmarklet ai preferiti del browser.\n"
                          "Quando sei su una pratica, cliccalo per inviare i dati all'app.",
                     font=ctk.CTkFont(size=FS()),
                     text_color=T_["testo"], justify="left").pack(anchor="w", pady=(0,10))

        for num, testo in [
            ("1","Copia il codice qui sotto"),
            ("2","Apri browser → mostra barra preferiti (Ctrl+Shift+B)"),
            ("3","Crea nuovo preferito → incolla il codice come URL"),
            ("4","Vai su una pratica e clicca il bookmark!"),
        ]:
            row = ctk.CTkFrame(body, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=num,
                         font=ctk.CTkFont(size=10, weight="bold"),
                         text_color="white", fg_color=T_["blu_medio"],
                         corner_radius=10, width=22, height=22).pack(side="left", padx=(0,8))
            ctk.CTkLabel(row, text=testo,
                         font=ctk.CTkFont(size=FS()),
                         text_color=T_["testo"], anchor="w").pack(side="left")

        ctk.CTkLabel(body, text="Codice:",
                     font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=T_["blu_scuro"]).pack(anchor="w", pady=(10,4))
        txt = ctk.CTkTextbox(body, height=72,
                             font=ctk.CTkFont(family="Consolas", size=9),
                             fg_color=T_["card"], border_color=T_["bordo"],
                             border_width=1, corner_radius=6)
        txt.pack(fill="x", pady=(0,8))
        txt.insert("1.0", BOOKMARKLET_JS)
        txt.configure(state="disabled")
        ctk.CTkButton(body, text="📋  Copia negli appunti",
                      command=lambda: [self.clipboard_clear(),
                                       self.clipboard_append(BOOKMARKLET_JS),
                                       messagebox.showinfo("Copiato","Codice copiato!")],
                      fg_color=T_["blu_medio"], hover_color=T_["blu_scuro"],
                      corner_radius=8, height=34,
                      font=ctk.CTkFont(size=12, weight="bold")
                      ).pack(fill="x")

# ─────────────────────────────────────────────
#  WIDGET RIUTILIZZABILI
# ─────────────────────────────────────────────
def Card(parent, **kw):
    T_ = T()
    return ctk.CTkFrame(parent, fg_color=T_["card"], corner_radius=10,
                        border_color=T_["bordo"], border_width=1, **kw)

def lbl_f(parent, text):
    T_ = T()
    return ctk.CTkLabel(parent, text=text, font=ctk.CTkFont(size=FS(-2)),
                        text_color=T_["testo_light"])

def campo_v(parent, text):
    T_ = T()
    val = str(text or "")
    e = ctk.CTkEntry(parent, font=ctk.CTkFont(size=FS()),
                     fg_color="transparent", border_width=0,
                     text_color=T_["testo"] if val else T_["testo_light"])
    e.insert(0, val if val else "—")
    e.configure(state="readonly")
    return e

# ─────────────────────────────────────────────
#  VISTA LISTA
# ─────────────────────────────────────────────
class VistaLista(ctk.CTkFrame):
    COLONNE = [
        ("Nr. Pratica",    "Nr. Pratica",    110),
        ("Cognome",        "Cognome",         120),
        ("Nome",           "Nome",            120),
        ("Data Nascita",   "Data Nascita",     95),
        ("Cittadinanza",   "Cittadinanza",    110),
        ("CUI",            "CUI",              85),
        ("Stato",          "_stato",          145),
        ("Data Estrazione","Data Estrazione", 130),
        ("Appuntamento",   "Data Appuntamento", 110),
    ]

    def __init__(self, parent, get_path, on_select, **kw):
        T_ = T()
        super().__init__(parent, fg_color=T_["sfondo"], corner_radius=0, **kw)
        self._get_path = get_path; self._on_select = on_select
        self._pratiche = []; self._stati = {}
        self._sort_col = "Nr. Pratica"; self._sort_asc = True
        self._build()

    def _build(self):
        T_ = T()
        bar1 = ctk.CTkFrame(self, fg_color=T_["card"], corner_radius=0, height=48)
        bar1.pack(fill="x"); bar1.pack_propagate(False)
        ctk.CTkLabel(bar1, text="Lista Pratiche",
                     font=ctk.CTkFont(size=FS(1), weight="bold"),
                     text_color=T_["blu_scuro"]).pack(side="left", padx=14)
        self._search_var = ctk.StringVar()
        self._search_var.trace_add("write", lambda *a: self._filtra())
        ctk.CTkEntry(bar1, textvariable=self._search_var,
                     placeholder_text="🔍  Cerca...",
                     fg_color=T_["sfondo"], border_color=T_["bordo"],
                     height=30, width=240, corner_radius=8,
                     font=ctk.CTkFont(size=FS())
                     ).pack(side="left", padx=(0,6))
        ctk.CTkButton(bar1, text="✕", command=lambda: self._search_var.set(""),
                      fg_color=T_["bordo"], text_color=T_["testo_light"],
                      hover_color=T_["sfondo"], height=30, width=32,
                      corner_radius=8).pack(side="left", padx=(0,8))
        ctk.CTkButton(bar1, text="↻  Aggiorna", command=self.carica,
                      fg_color=T_["blu_medio"], hover_color=T_["blu_scuro"],
                      height=30, width=90, corner_radius=8,
                      font=ctk.CTkFont(size=FS())).pack(side="left")
        self.lbl_count = ctk.CTkLabel(bar1, text="",
                                      font=ctk.CTkFont(size=FS()),
                                      text_color=T_["testo_light"])
        self.lbl_count.pack(side="right", padx=14)

        bar2 = ctk.CTkFrame(self, fg_color=T_["sfondo"], corner_radius=0, height=36)
        bar2.pack(fill="x"); bar2.pack_propagate(False)
        ctk.CTkLabel(bar2, text="Filtri:", font=ctk.CTkFont(size=FS()),
                     text_color=T_["testo_light"]).pack(side="left", padx=(14,6))
        self._stato_var = ctk.StringVar(value="Tutti gli stati")
        ctk.CTkOptionMenu(bar2, variable=self._stato_var,
                          values=["Tutti gli stati"]+[STATI[k][0] for k in STATI],
                          fg_color=T_["card"], button_color=T_["blu_medio"],
                          text_color=T_["testo"], font=ctk.CTkFont(size=FS()),
                          width=160, height=26, corner_radius=8,
                          command=lambda v: self._filtra()).pack(side="left", padx=(0,8))
        ctk.CTkLabel(bar2, text="Operatore:", font=ctk.CTkFont(size=FS()),
                     text_color=T_["testo_light"]).pack(side="left", padx=(0,4))
        self._op_var = ctk.StringVar(value="Tutti")
        self._op_menu = ctk.CTkOptionMenu(bar2, variable=self._op_var,
                          values=["Tutti"], fg_color=T_["card"],
                          button_color=T_["blu_medio"], text_color=T_["testo"],
                          font=ctk.CTkFont(size=FS()), width=130, height=26,
                          corner_radius=8, command=lambda v: self._filtra())
        self._op_menu.pack(side="left", padx=(0,8))
        ctk.CTkLabel(bar2, text="Scadenza:", font=ctk.CTkFont(size=FS()),
                     text_color=T_["testo_light"]).pack(side="left", padx=(0,4))
        self._scad_var = ctk.StringVar(value="Tutte")
        ctk.CTkOptionMenu(bar2, variable=self._scad_var,
                          values=["Tutte","30 giorni","60 giorni","90 giorni","6 mesi"],
                          fg_color=T_["card"], button_color=T_["blu_medio"],
                          text_color=T_["testo"], font=ctk.CTkFont(size=FS()),
                          width=110, height=26, corner_radius=8,
                          command=lambda v: self._filtra()).pack(side="left", padx=(0,8))
        ctk.CTkButton(bar2, text="↺ Reset", command=self._reset_filtri,
                      fg_color="transparent", text_color=T_["blu_medio"],
                      hover_color=T_["sfondo"], height=26, corner_radius=8,
                      font=ctk.CTkFont(size=FS())).pack(side="left")

        # ── Barra grafico stati ──
        self._bar_stati = ctk.CTkFrame(self, fg_color=T_["sfondo"],
                                       corner_radius=0, height=58)
        self._bar_stati.pack(fill="x", pady=(2,0))
        self._bar_stati.pack_propagate(False)
        # I widget vengono creati da _aggiorna_grafico_stati()

        frame = tk.Frame(self, bg=T_["sfondo"])
        frame.pack(fill="both", expand=True)
        vsb = tk.Scrollbar(frame, orient="vertical")
        vsb.pack(side="right", fill="y")
        hsb = tk.Scrollbar(frame, orient="horizontal")
        hsb.pack(side="bottom", fill="x")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Lista.Treeview",
                        background=T_["tree_bg"], foreground=T_["tree_fg"],
                        rowheight=26, fieldbackground=T_["tree_bg"],
                        borderwidth=0, font=("Segoe UI", FS()))
        style.configure("Lista.Treeview.Heading",
                        background=T_["blu_scuro"], foreground="white",
                        relief="flat", font=("Segoe UI", FS(), "bold"), padding=(8,6))
        style.map("Lista.Treeview",
                  background=[("selected", T_["tree_sel"])],
                  foreground=[("selected","white")])

        cols = [c[0] for c in self.COLONNE]
        self.tree = ttk.Treeview(frame, columns=cols, show="headings",
                                 style="Lista.Treeview",
                                 yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        for label,_,width in self.COLONNE:
            self.tree.heading(label, text=label, command=lambda c=label: self._ordina(c))
            self.tree.column(label, width=width, minwidth=50, anchor="w")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", self._apri)
        self.tree.bind("<Return>",   self._apri)

        self.tree.tag_configure("da_verificare", background="#FEF08A")
        self.tree.tag_configure("sospesa",       background="#FED7AA")
        self.tree.tag_configure("validata",      background="#BBF7D0")
        self.tree.tag_configure("negata",        background="#FECACA")
        self.tree.tag_configure("alt",           background="#F1F5F9")

        foot = ctk.CTkFrame(self, fg_color=T_["card"], corner_radius=0, height=26)
        foot.pack(fill="x"); foot.pack_propagate(False)
        ctk.CTkLabel(foot,
                     text="  Doppio click o Invio per aprire  •  Click intestazione per ordinare",
                     font=ctk.CTkFont(size=FS(-2)),
                     text_color=T_["testo_light"]).pack(side="left")

    def carica(self):
        path = self._get_path()
        self._pratiche = leggi_pratiche(path)
        self._stati    = leggi_tutti_stati(path)
        ops = sorted(set(str(p.get("Estratto Da","") or "").strip()
                         for p in self._pratiche if p.get("Estratto Da")))
        self._op_menu.configure(values=["Tutti"]+ops)
        self._filtra()

    def _reset_filtri(self):
        self._stato_var.set("Tutti gli stati")
        self._op_var.set("Tutti")
        self._scad_var.set("Tutte")
        self._search_var.set("")

    def _filtra(self):
        q     = self._search_var.get().lower()
        esiti = {STATI[k][0]: k for k in STATI}
        fkey  = esiti.get(self._stato_var.get(), "tutte")
        fop   = self._op_var.get()
        fscad = self._scad_var.get()
        oggi  = date.today()
        scad_gg = {"30 giorni":30,"60 giorni":60,"90 giorni":90,"6 mesi":180}
        out = []
        for p in self._pratiche:
            if q and not any(q in str(p.get(c,"")).lower()
                             for c in ["Cognome","Nome","CUI","Nr. Pratica","Cittadinanza"]):
                continue
            nr    = str(p.get("Nr. Pratica","")).strip()
            stato = self._stati.get(nr,"da_verificare")
            if fkey != "tutte" and stato != fkey: continue
            if fop  != "Tutti" and str(p.get("Estratto Da","")).strip() != fop: continue
            if fscad in scad_gg:
                try:
                    sd = datetime.strptime(str(p.get("Scadenza Rinnovo","")),"%d/%m/%Y").date()
                    if not (oggi <= sd <= oggi+timedelta(days=scad_gg[fscad])): continue
                except Exception: continue
            out.append((p, stato))

        col_map = {c[0]: c[1] for c in self.COLONNE}
        campo   = col_map.get(self._sort_col, self._sort_col)
        def sk(item):
            p, stato = item
            if campo == "_stato": return STATI.get(stato, STATI["da_verificare"])[0]
            return str(p.get(campo,"") or "").lower()
        out.sort(key=sk, reverse=not self._sort_asc)
        self._render(out)

    def _aggiorna_grafico_stati(self, righe):
        """Mostra card colorate con conteggio per stato."""
        T_ = T()
        for w in self._bar_stati.winfo_children():
            w.destroy()
        # Conta per stato
        conteggi = {k: 0 for k in STATI}
        for p, stato in righe:
            if stato in conteggi:
                conteggi[stato] += 1
        totale = len(righe)

        # Barra proporzionale
        bar_w = ctk.CTkFrame(self._bar_stati, fg_color="transparent")
        bar_w.pack(fill="x", padx=10, pady=(6,2))

        for stato_key, (label, bg_hex, fg_hex) in STATI.items():
            n = conteggi[stato_key]
            # Card cliccabile
            card = ctk.CTkFrame(bar_w,
                                fg_color=bg_hex,
                                corner_radius=8,
                                cursor="hand2")
            card.pack(side="left", padx=3, pady=0, expand=True, fill="x")
            # Numero grande
            ctk.CTkLabel(card,
                         text=str(n),
                         font=ctk.CTkFont(size=16, weight="bold"),
                         text_color=fg_hex,
                         fg_color="transparent"
                         ).pack(side="left", padx=(10,4), pady=4)
            # Etichetta
            lbl_short = label.split(" ",1)[1] if " " in label else label
            ctk.CTkLabel(card,
                         text=lbl_short,
                         font=ctk.CTkFont(size=10),
                         text_color=fg_hex,
                         fg_color="transparent"
                         ).pack(side="left", padx=(0,8), pady=4)
            # Click → filtra per quello stato
            for w in [card] + list(card.winfo_children()):
                w.bind("<Button-1>", lambda e, sk=stato_key: self._filtra_per_stato(sk))

        # Barra colorata proporzionale in fondo
        if totale > 0:
            barra = ctk.CTkFrame(self._bar_stati, fg_color=T_["bordo"],
                                 corner_radius=4, height=6)
            barra.pack(fill="x", padx=10, pady=(0,4))
            barra.update_idletasks()
            bw = barra.winfo_width() or 400
            x = 0
            for stato_key, (_, bg_hex, fg_hex) in STATI.items():
                n = conteggi[stato_key]
                if n == 0: continue
                w_seg = int(bw * n / totale)
                seg = tk.Frame(barra, bg=fg_hex, height=6, width=w_seg)
                seg.place(x=x, y=0)
                x += w_seg

    def _filtra_per_stato(self, stato_key: str):
        """Filtra la lista per lo stato cliccato (toggle)."""
        etichetta = STATI[stato_key][0]
        if self._stato_var.get() == etichetta:
            self._stato_var.set("Tutti gli stati")
        else:
            self._stato_var.set(etichetta)
        self._filtra()

    def _render(self, righe):
        self.tree.delete(*self.tree.get_children())
        self.lbl_count.configure(text=f"{len(righe)} pratiche")
        self._aggiorna_grafico_stati(righe)
        for i,(p,stato) in enumerate(righe):
            nr = str(p.get("Nr. Pratica",""))
            valori = (nr,
                      str(p.get("Cognome","") or ""),
                      str(p.get("Nome","") or ""),
                      str(p.get("Data Nascita","") or ""),
                      str(p.get("Cittadinanza","") or p.get("Nazione Nascita","") or ""),
                      str(p.get("CUI","") or ""),
                      STATI.get(stato, STATI["da_verificare"])[0],
                      str(p.get("Data Estrazione","") or ""),
                      str(p.get("Data Appuntamento","") or ""))
            tag = stato if stato in STATI else ("alt" if i%2 else "")
            self.tree.insert("","end", iid=nr, values=valori, tags=(tag,))

    def _ordina(self, col):
        if self._sort_col == col: self._sort_asc = not self._sort_asc
        else: self._sort_col = col; self._sort_asc = True
        self._filtra()

    def _apri(self, event=None):
        sel = self.tree.selection()
        if sel: self._on_select(sel[0])

# ─────────────────────────────────────────────
#  CALENDARIO POPUP
# ─────────────────────────────────────────────
class CalendarioPopup(ctk.CTkToplevel):
    """Piccolo calendario per selezionare una data."""
    def __init__(self, parent, data_var: ctk.StringVar, anchor_widget):
        super().__init__(parent)
        T_ = T()
        self.overrideredirect(True)   # niente bordi finestra
        self.configure(fg_color=T_["card"])
        self._data_var = data_var
        self._build()
        # Posiziona vicino al widget di ancoraggio
        self.update_idletasks()
        x = anchor_widget.winfo_rootx()
        y = anchor_widget.winfo_rooty() + anchor_widget.winfo_height() + 2
        self.geometry(f"+{x}+{y}")
        self.lift()
        self.focus_force()
        self.bind("<FocusOut>", lambda e: self.after(100, self._check_focus))

    def _check_focus(self):
        try:
            if not self.focus_get():
                self.destroy()
        except Exception:
            self.destroy()

    def _build(self):
        T_ = T()
        oggi = date.today()
        # Usa la data corrente nel campo se valida, altrimenti oggi
        try:
            self._current = datetime.strptime(
                self._data_var.get().strip(), "%d/%m/%Y").date()
        except Exception:
            self._current = oggi

        self._anno  = self._current.year
        self._mese  = self._current.month
        self._render()

    def _render(self):
        T_ = T()
        for w in self.winfo_children():
            w.destroy()

        # Bordo esterno
        outer = ctk.CTkFrame(self, fg_color=T_["card"], corner_radius=10,
                             border_color=T_["bordo"], border_width=1)
        outer.pack(padx=2, pady=2)

        # Header mese/anno
        hdr = ctk.CTkFrame(outer, fg_color=T_["blu_scuro"], corner_radius=8)
        hdr.pack(fill="x", padx=6, pady=(6,4))

        ctk.CTkButton(hdr, text="◀", width=28, height=26,
                      fg_color="transparent", hover_color=T_["blu_medio"],
                      text_color="white", font=ctk.CTkFont(size=12),
                      command=self._mese_prec).pack(side="left", padx=2)

        import calendar
        nome_mese = ["","Gennaio","Febbraio","Marzo","Aprile","Maggio","Giugno",
                     "Luglio","Agosto","Settembre","Ottobre","Novembre","Dicembre"]
        ctk.CTkLabel(hdr, text=f"{nome_mese[self._mese]} {self._anno}",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color="white").pack(side="left", expand=True)

        ctk.CTkButton(hdr, text="▶", width=28, height=26,
                      fg_color="transparent", hover_color=T_["blu_medio"],
                      text_color="white", font=ctk.CTkFont(size=12),
                      command=self._mese_succ).pack(side="right", padx=2)

        # Giorni settimana
        giorni_hdr = ctk.CTkFrame(outer, fg_color="transparent")
        giorni_hdr.pack(padx=6)
        for g in ["Lu","Ma","Me","Gi","Ve","Sa","Do"]:
            ctk.CTkLabel(giorni_hdr, text=g, width=32, height=22,
                         font=ctk.CTkFont(size=10),
                         text_color=T_["testo_light"]).pack(side="left")

        # Griglia giorni
        griglia = ctk.CTkFrame(outer, fg_color="transparent")
        griglia.pack(padx=6, pady=(0,6))

        import calendar as cal
        primo_giorno, num_giorni = cal.monthrange(self._anno, self._mese)
        oggi = date.today()

        # Celle vuote prima del primo giorno
        riga = ctk.CTkFrame(griglia, fg_color="transparent")
        riga.pack()
        for _ in range(primo_giorno):
            ctk.CTkLabel(riga, text="", width=32, height=28).pack(side="left")

        col = primo_giorno
        for giorno in range(1, num_giorni + 1):
            d = date(self._anno, self._mese, giorno)
            is_oggi     = d == oggi
            is_selected = d == self._current

            if is_selected:
                fg = T_["blu_medio"]; tc = "white"
            elif is_oggi:
                fg = T_["blu_chiaro"] if "blu_chiaro" in T_ else "#D6E4F0"; tc = T_["blu_scuro"]
            elif d.weekday() >= 5:
                fg = "transparent"; tc = T_["arancio"]
            else:
                fg = "transparent"; tc = T_["testo"]

            btn = ctk.CTkButton(riga, text=str(giorno),
                                width=32, height=28,
                                fg_color=fg,
                                hover_color=T_["blu_chiaro"] if "blu_chiaro" in T_ else "#D6E4F0",
                                text_color=tc,
                                corner_radius=6,
                                font=ctk.CTkFont(size=11),
                                command=lambda d=d: self._seleziona(d))
            btn.pack(side="left")
            col += 1
            if col % 7 == 0:
                riga = ctk.CTkFrame(griglia, fg_color="transparent")
                riga.pack()

        # Bottone Oggi
        ctk.CTkButton(outer, text="Oggi",
                      command=lambda: self._seleziona(date.today()),
                      fg_color=T_["sfondo"], hover_color=T_["blu_medio"],
                      text_color=T_["testo_light"], height=24,
                      corner_radius=6, font=ctk.CTkFont(size=10)
                      ).pack(fill="x", padx=6, pady=(0,6))

    def _mese_prec(self):
        if self._mese == 1:
            self._mese = 12; self._anno -= 1
        else:
            self._mese -= 1
        self._render()

    def _mese_succ(self):
        if self._mese == 12:
            self._mese = 1; self._anno += 1
        else:
            self._mese += 1
        self._render()

    def _seleziona(self, d: date):
        self._data_var.set(d.strftime("%d/%m/%Y"))
        self.destroy()


# ─────────────────────────────────────────────
#  FINESTRA TRAINANTE (affiancata)
# ─────────────────────────────────────────────
class FinestraTrainante(ctk.CTkToplevel):
    def __init__(self, parent, nr_trainante: str, get_path):
        super().__init__(parent)
        T_ = T()
        self.title(f"Pratica Trainante — {nr_trainante}")
        self.geometry("680x700")
        self.configure(fg_color=T_["sfondo"])
        self.lift()
        self._get_path    = get_path
        self._nr          = nr_trainante
        self._stato_var   = None
        self._stato_menu  = None
        self._appuntamento_var = ctk.StringVar(value="")
        self._build(nr_trainante)

    def _build(self, nr):
        T_ = T()
        pratiche = leggi_pratiche(self._get_path())
        pratica  = next((p for p in pratiche
                         if str(p.get("Nr. Pratica","")).strip() == nr), None)
        if not pratica:
            ctk.CTkLabel(self,
                         text=f"Pratica {nr} non trovata. Estraila dal gestionale.",
                         font=ctk.CTkFont(size=13),
                         text_color=T_["arancio"],
                         justify="center").pack(expand=True)
            return

        att = leggi_attivita(self._get_path(), nr)
        pratica["Stato"] = att.get("Stato","da_verificare") or "da_verificare"

        # Header
        hdr = ctk.CTkFrame(self, fg_color=T_["blu_scuro"],
                           corner_radius=0, height=52)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        nome = f"{pratica.get('Cognome','')} {pratica.get('Nome','')}".strip()
        ctk.CTkLabel(hdr, text=f"  🔗  Trainante: {nome}",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color="white").pack(side="left", padx=14)
        ctk.CTkLabel(hdr, text=nr,
                     font=ctk.CTkFont(family="Consolas", size=12),
                     text_color="#93C5FD").pack(side="right", padx=14)

        scroll = ctk.CTkScrollableFrame(self, fg_color=T_["sfondo"], corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=12, pady=10)

        # Nr + Stato
        nr_card = ctk.CTkFrame(scroll, fg_color=T_["card"], corner_radius=10,
                               border_color=T_["bordo"], border_width=1)
        nr_card.pack(fill="x", pady=(0,8))
        ni = ctk.CTkFrame(nr_card, fg_color="transparent")
        ni.pack(fill="x", padx=14, pady=10)
        ctk.CTkLabel(ni, text="Nr. Pratica", font=ctk.CTkFont(size=10),
                     text_color=T_["testo_light"]).pack(anchor="w")
        ctk.CTkLabel(ni, text=nr,
                     font=ctk.CTkFont(family="Consolas", size=20, weight="bold"),
                     text_color=T_["blu_scuro"]).pack(anchor="w")
        stato_key  = str(pratica.get("Stato","da_verificare") or "da_verificare")
        info_stato = STATI.get(stato_key, STATI["da_verificare"])
        ctk.CTkLabel(ni, text=info_stato[0],
                     font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=info_stato[2],
                     fg_color=info_stato[1],
                     corner_radius=16).pack(anchor="w", pady=(6,0))

        # Anagrafica compatta
        ana = ctk.CTkFrame(scroll, fg_color=T_["card"], corner_radius=10,
                           border_color=T_["bordo"], border_width=1)
        ana.pack(fill="x", pady=(0,8))
        ctk.CTkLabel(ana, text="👤  Anagrafica",
                     font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=T_["blu_scuro"]).pack(anchor="w", padx=14, pady=(10,4))
        ctk.CTkFrame(ana, fg_color=T_["bordo"], height=1, corner_radius=0
                     ).pack(fill="x", padx=14, pady=(0,8))

        campi = [
            ("Cognome",         pratica.get("Cognome")),
            ("Nome",            pratica.get("Nome")),
            ("CUI",             pratica.get("CUI")),
            ("Codice Fiscale",  pratica.get("Codice Fiscale")),
            ("Data Nascita",    pratica.get("Data Nascita")),
            ("Cittadinanza",    pratica.get("Cittadinanza")),
            ("Telefono",        pratica.get("Telefono")),
            ("Tipo Soggiorno",  pratica.get("Tipo Soggiorno")),
            ("Motivo Soggiorno",pratica.get("Motivo Soggiorno")),
            ("Scadenza Rinnovo",pratica.get("Scadenza Rinnovo")),
            ("Comune",          pratica.get("Comune")),
            ("Indirizzo",       pratica.get("Indirizzo")),
        ]
        g = ctk.CTkFrame(ana, fg_color="transparent")
        g.pack(fill="x", padx=14, pady=(0,12))
        g.columnconfigure(1, weight=1); g.columnconfigure(3, weight=1)
        for i,(lbl,val) in enumerate(campi):
            r,cl = i//2,(i%2)*2
            ctk.CTkLabel(g, text=lbl, font=ctk.CTkFont(size=FS(-2)),
                         text_color=T_["testo_light"]
                         ).grid(row=r, column=cl, sticky="nw", padx=(0,6), pady=2)
            e = ctk.CTkEntry(g, font=ctk.CTkFont(size=FS()),
                             fg_color="transparent", border_width=0,
                             text_color=T_["testo"])
            e.insert(0, str(val or "—"))
            e.configure(state="readonly")
            e.grid(row=r, column=cl+1, sticky="ew", pady=2, padx=(0,16))

        # Note
        if att.get("Note Personali"):
            nc = ctk.CTkFrame(scroll, fg_color=T_["card"], corner_radius=10,
                              border_color=T_["bordo"], border_width=1)
            nc.pack(fill="x", pady=(0,8))
            ctk.CTkLabel(nc, text="📝  Note",
                         font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=T_["blu_scuro"]).pack(anchor="w", padx=14, pady=(10,4))
            ctk.CTkLabel(nc, text=str(att.get("Note Personali","")),
                         font=ctk.CTkFont(size=FS()), text_color=T_["testo"],
                         wraplength=580, justify="left"
                         ).pack(anchor="w", padx=14, pady=(0,12))


# ─────────────────────────────────────────────
#  VISTA SCHEDA
# ─────────────────────────────────────────────
class VistaScheda(ctk.CTkFrame):
    def __init__(self, parent, get_path, vai_lista, appuntamento_var=None, **kw):
        T_ = T()
        super().__init__(parent, fg_color=T_["sfondo"], corner_radius=0, **kw)
        self._get_path  = get_path
        self._vai_lista = vai_lista
        self._nr = None; self._stato_var = None; self._stato_menu = None
        self._pratica_corrente = None; self._att_corrente = None
        self._appuntamento_var = appuntamento_var  # variabile persistente dall'App
        self._finestra_trainante = None
        self._build()

    def _build(self):
        T_ = T()
        nav = ctk.CTkFrame(self, fg_color=T_["nav_bg"], corner_radius=0, height=44)
        nav.pack(fill="x"); nav.pack_propagate(False)
        ctk.CTkButton(nav, text="☰  Lista Pratiche", command=self._vai_lista,
                      fg_color="transparent", text_color=T_["blu_medio"],
                      hover_color=T_["sfondo"], height=32, corner_radius=8,
                      font=ctk.CTkFont(size=FS())).pack(side="left", padx=(8,0))
        ctk.CTkLabel(nav, text="›", font=ctk.CTkFont(size=14),
                     text_color=T_["testo_light"]).pack(side="left", padx=4)
        self.lbl_breadcrumb = ctk.CTkLabel(nav, text="—",
                                           font=ctk.CTkFont(size=FS(), weight="bold"),
                                           text_color=T_["testo"])
        self.lbl_breadcrumb.pack(side="left")

        # Campo data appuntamento (a destra nella nav)
        ctk.CTkLabel(nav, text="📅 Appuntamento:",
                     font=ctk.CTkFont(size=FS()),
                     text_color=T_["testo_light"]).pack(side="right", padx=(0,4))
        # Usa la variabile persistente dell'App se disponibile
        if not self._appuntamento_var:
            self._appuntamento_var = ctk.StringVar(value="")
        self._appuntamento_entry = ctk.CTkEntry(nav,
                                                textvariable=self._appuntamento_var,
                                                placeholder_text="GG/MM/AAAA",
                                                fg_color=T_["sfondo"],
                                                border_color=T_["bordo"],
                                                height=28, width=110,
                                                corner_radius=6,
                                                font=ctk.CTkFont(size=FS()))
        self._appuntamento_entry.pack(side="right", padx=(0,4))

        # Bottone calendario
        ctk.CTkButton(nav, text="🗓",
                      command=lambda: CalendarioPopup(
                          self, self._appuntamento_var, self._appuntamento_entry),
                      fg_color=T_["blu_medio"], hover_color=T_["blu_scuro"],
                      width=30, height=28, corner_radius=6,
                      font=ctk.CTkFont(size=13)).pack(side="right", padx=(0,2))

        # Bottone salva
        ctk.CTkButton(nav, text="💾",
                      command=self._salva_appuntamento,
                      fg_color=T_["verde"], hover_color=T_["verde_dark"],
                      width=30, height=28, corner_radius=6,
                      font=ctk.CTkFont(size=11)).pack(side="right", padx=(0,8))



        self._corpo = ctk.CTkFrame(self, fg_color=T_["sfondo"], corner_radius=0)
        self._corpo.pack(fill="both", expand=True)
        self._frame_scheda = ctk.CTkFrame(self._corpo, fg_color=T_["sfondo"], corner_radius=0)
        self._frame_scheda.pack(side="left", fill="both", expand=True)
        ctk.CTkFrame(self._corpo, fg_color=T_["bordo"], width=1, corner_radius=0).pack(side="left", fill="y")
        self._frame_att = ctk.CTkFrame(self._corpo, fg_color=T_["sfondo"],
                                       corner_radius=0, width=340)
        self._frame_att.pack(side="left", fill="y"); self._frame_att.pack_propagate(False)
        ctk.CTkLabel(self._frame_scheda, text="← Seleziona una pratica",
                     font=ctk.CTkFont(size=13), text_color=T_["testo_light"]).pack(expand=True)

    def carica(self, nr):
        self._nr = nr
        path = self._get_path()
        pratiche = leggi_pratiche(path)
        pratica = next((p for p in pratiche if str(p.get("Nr. Pratica","")).strip()==nr), None)
        if not pratica: return
        att = leggi_attivita(path, nr)
        pratica["Stato"] = att.get("Stato","da_verificare") or "da_verificare"
        self._pratica_corrente = pratica
        self._att_corrente     = att
        nome = f"{pratica.get('Cognome','')} {pratica.get('Nome','')}".strip()
        self.lbl_breadcrumb.configure(text=f"{nr}  —  {nome}")
        # Carica data appuntamento dalla pratica SOLO se il campo è vuoto
        if self._appuntamento_var:
            data_pratica = str(pratica.get("Data Appuntamento","") or "")
            if not self._appuntamento_var.get().strip() and data_pratica:
                self._appuntamento_var.set(data_pratica)
        self._render_scheda(pratica, att)
        self._render_attivita(nr, att)

    def _apri_trainante(self, nr_trainante: str):
        if not nr_trainante or nr_trainante == "—":
            return
        # Chiudi finestra precedente se aperta
        if self._finestra_trainante and self._finestra_trainante.winfo_exists():
            self._finestra_trainante.destroy()
        self._finestra_trainante = FinestraTrainante(
            self.winfo_toplevel(), nr_trainante, self._get_path)

    def _salva_appuntamento(self):
        if not self._nr or not self._pratica_corrente:
            return
        data = self._appuntamento_var.get().strip()
        # Validazione formato data
        if data:
            try:
                datetime.strptime(data, "%d/%m/%Y")
            except ValueError:
                messagebox.showwarning("Formato errato",
                                       "Inserisci la data nel formato GG/MM/AAAA")
                return
        self._pratica_corrente["Data Appuntamento"] = data
        salva_pratica_excel(self._get_path(), self._pratica_corrente)
        messagebox.showinfo("Salvato",
                            f"Appuntamento salvato: {data if data else '(rimosso)'}")

    def aggiorna_dati(self, dati_nuovi: dict):
        """Aggiorna silenziosamente la pratica corrente senza popup."""
        if self._nr and self._nr == dati_nuovi.get("Nr. Pratica","").strip():
            # Preserva la data appuntamento già impostata
            if self._appuntamento_var:
                data_app = self._appuntamento_var.get().strip()
                if data_app:
                    dati_nuovi["Data Appuntamento"] = data_app
            salva_pratica_excel(self._get_path(), dati_nuovi)
            self.carica(self._nr)
            return True
        return False

    def _render_scheda(self, pratica, att):
        T_ = T()
        for w in self._frame_scheda.winfo_children(): w.destroy()
        scroll = ctk.CTkScrollableFrame(self._frame_scheda, fg_color=T_["sfondo"], corner_radius=0)
        scroll.pack(fill="both", expand=True, padx=14, pady=12)
        nr = str(pratica.get("Nr. Pratica","") or "")

        nr_card = Card(scroll); nr_card.pack(fill="x", pady=(0,8))
        ni = ctk.CTkFrame(nr_card, fg_color="transparent"); ni.pack(fill="x", padx=14, pady=12)
        ctk.CTkLabel(ni, text="Nr. Pratica", font=ctk.CTkFont(size=FS(-2)),
                     text_color=T_["testo_light"]).pack(anchor="w")
        ctk.CTkLabel(ni, text=nr,
                     font=ctk.CTkFont(family="Consolas", size=22, weight="bold"),
                     text_color=T_["blu_scuro"]).pack(anchor="w")

        stato_key  = str(pratica.get("Stato","da_verificare") or "da_verificare")
        info_stato = STATI.get(stato_key, STATI["da_verificare"])
        self._stato_var = ctk.StringVar(value=info_stato[0])
        esiti_map = {STATI[k][0]: k for k in STATI}
        sr = ctk.CTkFrame(ni, fg_color="transparent"); sr.pack(anchor="w", fill="x", pady=(8,0))
        self._stato_menu = ctk.CTkOptionMenu(sr, variable=self._stato_var,
                           values=[STATI[k][0] for k in STATI],
                           fg_color=info_stato[1], button_color=info_stato[2],
                           text_color=info_stato[2],
                           font=ctk.CTkFont(size=FS(), weight="bold"),
                           height=30, width=175, corner_radius=16,
                           command=lambda v: self._aggiorna_colore(esiti_map.get(v,"da_verificare")))
        self._stato_menu.pack(side="left", padx=(0,10))
        ctk.CTkButton(sr, text="💾 Salva stato",
                      command=lambda: self._salva_stato(nr, esiti_map),
                      fg_color=T_["verde"], hover_color=T_["verde_dark"],
                      height=30, corner_radius=8, font=ctk.CTkFont(size=FS())
                      ).pack(side="left")
        ctk.CTkLabel(nr_card,
                     text=f"Estratto da {pratica.get('Estratto Da','')}  •  {pratica.get('Data Estrazione','')}",
                     font=ctk.CTkFont(size=9), text_color=T_["testo_light"]
                     ).pack(anchor="e", padx=14, pady=(0,8))

        ana = Card(scroll); ana.pack(fill="x", pady=(0,8))
        ctk.CTkLabel(ana, text="👤  Anagrafica",
                     font=ctk.CTkFont(size=FS(), weight="bold"),
                     text_color=T_["blu_scuro"]).pack(anchor="w", padx=14, pady=(10,4))
        ctk.CTkFrame(ana, fg_color=T_["bordo"], height=1, corner_radius=0).pack(fill="x", padx=14, pady=(0,8))
        campi = [
            ("Cognome",pratica.get("Cognome")),("Nome",pratica.get("Nome")),
            ("CUI",pratica.get("CUI")),("Codice Fiscale",pratica.get("Codice Fiscale")),
            ("Data Nascita",pratica.get("Data Nascita")),("Luogo Nascita",pratica.get("Luogo Nascita")),
            ("Nazione Nascita",pratica.get("Nazione Nascita")),("Cittadinanza",pratica.get("Cittadinanza")),
            ("Sesso",pratica.get("Sesso")),("Stato Civile",pratica.get("Stato Civile")),
            ("Telefono",pratica.get("Telefono")),("Coniuge",pratica.get("Coniuge")),
            ("Comune",pratica.get("Comune")),("Indirizzo",pratica.get("Indirizzo")),
            ("Tipo Soggiorno",pratica.get("Tipo Soggiorno")),("Tipo Pratica",pratica.get("Tipo Pratica")),
            ("Motivo Soggiorno",pratica.get("Motivo Soggiorno")),("Validita Sogg.",pratica.get("Validita Soggiorno")),
            ("Presentazione",pratica.get("Data Presentazione")),("Scadenza Rinnovo",pratica.get("Scadenza Rinnovo")),
            ("Referenze",pratica.get("Referenze")),("Op. Gestionale",pratica.get("Operatore Gestionale")),
        ]
        g = ctk.CTkFrame(ana, fg_color="transparent"); g.pack(fill="x", padx=14, pady=(0,12))
        g.columnconfigure(1, weight=1); g.columnconfigure(3, weight=1)
        for i,(lbl,val) in enumerate(campi):
            r,cl = i//2,(i%2)*2
            lbl_f(g,lbl).grid(row=r,column=cl,sticky="nw",padx=(0,6),pady=2)
            campo_v(g,val).grid(row=r,column=cl+1,sticky="ew",pady=2,padx=(0,16))

        if pratica.get("Note Gestionale"):
            nc = Card(scroll); nc.pack(fill="x", pady=(0,8))
            ctk.CTkLabel(nc, text="📝  Note gestionale",
                         font=ctk.CTkFont(size=FS(), weight="bold"),
                         text_color=T_["blu_scuro"]).pack(anchor="w", padx=14, pady=(10,4))
            ctk.CTkFrame(nc, fg_color=T_["bordo"], height=1, corner_radius=0).pack(fill="x", padx=14, pady=(0,6))
            ctk.CTkLabel(nc, text=str(pratica.get("Note Gestionale","")),
                         font=ctk.CTkFont(size=FS()), text_color=T_["testo"],
                         wraplength=380, justify="left").pack(anchor="w", padx=14, pady=(0,12))

    def _render_attivita(self, nr, att):
        T_ = T()
        for w in self._frame_att.winfo_children(): w.destroy()
        hdr = ctk.CTkFrame(self._frame_att, fg_color=T_["blu_scuro"], corner_radius=0, height=44)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text="Attività",
                     font=ctk.CTkFont(size=FS(1), weight="bold"),
                     text_color="white").pack(side="left", padx=14, pady=10)
        tab = ctk.CTkTabview(self._frame_att, fg_color=T_["sfondo"],
                             segmented_button_fg_color=T_["card"],
                             segmented_button_selected_color=T_["blu_medio"],
                             segmented_button_selected_hover_color=T_["blu_scuro"],
                             segmented_button_unselected_color=T_["card"],
                             text_color=T_["testo"], corner_radius=0)
        tab.pack(fill="both", expand=True)
        tab.add("📝 Note"); tab.add("✅ Checklist")
        self._build_note(tab.tab("📝 Note"), nr, att)
        self._build_checklist(tab.tab("✅ Checklist"), nr, att)

    def _build_note(self, parent, nr, att):
        T_ = T()
        scroll = ctk.CTkScrollableFrame(parent, fg_color=T_["sfondo"], corner_radius=0)
        scroll.pack(fill="both", expand=True)
        nc = Card(scroll); nc.pack(fill="x", padx=6, pady=(6,4))
        ctk.CTkLabel(nc, text="📝  La mia nota",
                     font=ctk.CTkFont(size=FS(), weight="bold"),
                     text_color=T_["blu_scuro"]).pack(anchor="w", padx=10, pady=(8,4))
        self._nota_txt = ctk.CTkTextbox(nc, height=240,
                                        font=ctk.CTkFont(size=FS()),
                                        fg_color=T_["sfondo"], border_color=T_["bordo"],
                                        border_width=1, corner_radius=6)
        self._nota_txt.pack(fill="x", padx=10, pady=(0,8))
        self._nota_txt.insert("1.0", str(att.get("Note Personali","") or ""))
        ctk.CTkButton(nc, text="💾  Salva nota",
                      command=lambda: self._salva_nota(nr),
                      fg_color=T_["blu_medio"], hover_color=T_["blu_scuro"],
                      corner_radius=8, height=32, font=ctk.CTkFont(size=FS())
                      ).pack(anchor="e", padx=10, pady=(0,10))
        if att.get("Note Modificate Il"):
            ctk.CTkLabel(scroll, text=f"Ultima modifica: {att['Note Modificate Il']}",
                         font=ctk.CTkFont(size=9), text_color=T_["testo_light"]).pack(pady=2)

    def _build_checklist(self, parent, nr, att):
        T_ = T()
        scroll = ctk.CTkScrollableFrame(parent, fg_color=T_["sfondo"], corner_radius=0)
        scroll.pack(fill="both", expand=True)
        def sezione(titolo):
            c = Card(scroll); c.pack(fill="x", padx=6, pady=(6,3))
            ctk.CTkLabel(c, text=titolo, font=ctk.CTkFont(size=FS(), weight="bold"),
                         text_color=T_["blu_scuro"]).pack(anchor="w", padx=10, pady=(8,4))
            return c

        s1 = sezione("SDI")
        lbl_f(s1,"Esito").pack(anchor="w", padx=10)
        self._sdi_esito = ctk.CTkOptionMenu(s1, values=SDI_ESITI,
                           fg_color=T_["sfondo"], button_color=T_["blu_medio"],
                           text_color=T_["testo"], font=ctk.CTkFont(size=FS()),
                           height=28, corner_radius=6)
        self._sdi_esito.set(str(att.get("SDI Esito","") or ""))
        self._sdi_esito.pack(fill="x", padx=10, pady=(2,6))
        lbl_f(s1,"Note SDI").pack(anchor="w", padx=10)
        self._sdi_note = ctk.CTkTextbox(s1, height=52, font=ctk.CTkFont(size=FS()),
                                        fg_color=T_["sfondo"], border_color=T_["bordo"],
                                        border_width=1, corner_radius=6)
        self._sdi_note.pack(fill="x", padx=10, pady=(2,10))
        self._sdi_note.insert("1.0", str(att.get("SDI Note","") or ""))

        s2 = sezione("Reddito")
        lbl_f(s2,"Tipo").pack(anchor="w", padx=10)
        self._reddito_tipo = ctk.CTkOptionMenu(s2, values=REDDITO_TIPI,
                              fg_color=T_["sfondo"], button_color=T_["blu_medio"],
                              text_color=T_["testo"], font=ctk.CTkFont(size=FS()),
                              height=28, corner_radius=6)
        self._reddito_tipo.set(str(att.get("Reddito Tipo","") or ""))
        self._reddito_tipo.pack(fill="x", padx=10, pady=(2,6))
        lbl_f(s2,"Note reddito").pack(anchor="w", padx=10)
        self._reddito_note = ctk.CTkTextbox(s2, height=52, font=ctk.CTkFont(size=FS()),
                                            fg_color=T_["sfondo"], border_color=T_["bordo"],
                                            border_width=1, corner_radius=6)
        self._reddito_note.pack(fill="x", padx=10, pady=(2,10))
        self._reddito_note.insert("1.0", str(att.get("Reddito Note","") or ""))

        s3 = sezione("Famiglia / Trainante")
        self._fam_var = ctk.BooleanVar(value=str(att.get("Famiglia Trainante","")).upper()=="SI")
        ctk.CTkCheckBox(s3, text="Presenza trainante", variable=self._fam_var,
                        fg_color=T_["blu_medio"], hover_color=T_["blu_scuro"],
                        font=ctk.CTkFont(size=FS()), command=self._toggle_fam
                        ).pack(anchor="w", padx=10, pady=(0,6))
        self._fam_frame = ctk.CTkFrame(s3, fg_color="transparent")
        self._fam_frame.pack(fill="x", padx=10, pady=(0,10))
        self._fam_fields = {}
        for label,key in [("Nome trainante","Nome Trainante"),
                           ("CF trainante","CF Trainante"),
                           ("Nr. pratica trainante","Nr Pratica Trainante")]:
            lbl_f(self._fam_frame, label).pack(anchor="w")
            if key == "Nr Pratica Trainante":
                # Riga con entry + bottone link
                row_t = ctk.CTkFrame(self._fam_frame, fg_color="transparent")
                row_t.pack(fill="x", pady=(2,4))
                e = ctk.CTkEntry(row_t, fg_color=T_["sfondo"],
                                 border_color=T_["bordo"], height=28, corner_radius=6,
                                 font=ctk.CTkFont(size=FS()))
                e.pack(side="left", fill="x", expand=True, padx=(0,4))
                e.insert(0, str(att.get(key,"") or ""))
                ctk.CTkButton(row_t, text="🔗",
                              command=lambda e=e: self._apri_trainante(e.get().strip()),
                              fg_color=T_["blu_medio"], hover_color=T_["blu_scuro"],
                              width=30, height=28, corner_radius=6,
                              font=ctk.CTkFont(size=12)).pack(side="left")
            else:
                e = ctk.CTkEntry(self._fam_frame, fg_color=T_["sfondo"],
                                 border_color=T_["bordo"], height=28, corner_radius=6,
                                 font=ctk.CTkFont(size=FS()))
                e.pack(fill="x", pady=(2,4))
                e.insert(0, str(att.get(key,"") or ""))
            self._fam_fields[key] = e
        self._toggle_fam()

        s4 = sezione("Documenti mancanti")
        self._doc = ctk.CTkTextbox(s4, height=68, font=ctk.CTkFont(size=FS()),
                                   fg_color=T_["sfondo"], border_color=T_["bordo"],
                                   border_width=1, corner_radius=6)
        self._doc.pack(fill="x", padx=10, pady=(0,10))
        self._doc.insert("1.0", str(att.get("Documenti Mancanti","") or ""))

        ctk.CTkButton(scroll, text="💾  Salva Checklist",
                      command=lambda: self._salva_checklist(nr),
                      fg_color=T_["verde"], hover_color=T_["verde_dark"],
                      corner_radius=8, height=34,
                      font=ctk.CTkFont(size=FS(), weight="bold")
                      ).pack(fill="x", padx=6, pady=8)

    def _toggle_fam(self):
        T_ = T()
        st = "normal" if self._fam_var.get() else "disabled"
        for e in self._fam_fields.values():
            e.configure(state=st, fg_color=T_["sfondo"] if st=="normal" else T_["bordo"])

    def _aggiorna_colore(self, stato_key):
        info = STATI.get(stato_key, STATI["da_verificare"])
        self._stato_menu.configure(fg_color=info[1], button_color=info[2], text_color=info[2])

    def _salva_stato(self, nr, esiti_map):
        nuovo = esiti_map.get(self._stato_var.get(),"da_verificare")
        att = leggi_attivita(self._get_path(), nr)
        att["Nr. Pratica"] = nr; att["Stato"] = nuovo
        if not att.get("SDI Esito"): att["SDI Esito"] = ""
        salva_attivita_excel(self._get_path(), att)
        messagebox.showinfo("Salvato", f"Stato: {STATI[nuovo][0]}")

    def _salva_nota(self, nr):
        att = leggi_attivita(self._get_path(), nr)
        att.update({"Nr. Pratica":nr,
                    "Note Personali":self._nota_txt.get("1.0","end").strip(),
                    "Note Modificate Il":datetime.now().strftime("%d/%m/%Y %H:%M")})
        if not att.get("Stato"): att["Stato"] = "da_verificare"
        salva_attivita_excel(self._get_path(), att)
        messagebox.showinfo("Salvato","Nota salvata.")

    def _salva_checklist(self, nr):
        att = leggi_attivita(self._get_path(), nr)
        att.update({"Nr. Pratica":nr,
                    "SDI Esito":self._sdi_esito.get(),
                    "SDI Note":self._sdi_note.get("1.0","end").strip(),
                    "Reddito Tipo":self._reddito_tipo.get(),
                    "Reddito Note":self._reddito_note.get("1.0","end").strip(),
                    "Famiglia Trainante":"SI" if self._fam_var.get() else "NO",
                    "Nome Trainante":self._fam_fields["Nome Trainante"].get(),
                    "CF Trainante":self._fam_fields["CF Trainante"].get(),
                    "Nr Pratica Trainante":self._fam_fields["Nr Pratica Trainante"].get(),
                    "Documenti Mancanti":self._doc.get("1.0","end").strip(),
                    "Checklist Compilata Da":OPERATORE,
                    "Checklist Compilata Il":datetime.now().strftime("%d/%m/%Y %H:%M")})
        if not att.get("Stato"): att["Stato"] = "da_verificare"
        salva_attivita_excel(self._get_path(), att)
        messagebox.showinfo("Salvato","Checklist salvata.")

    def _stampa_pdf(self):
        if not self._pratica_corrente:
            messagebox.showwarning("Attenzione","Nessuna pratica selezionata.")
            return
        nr  = str(self._pratica_corrente.get("Nr. Pratica",""))
        default = f"Pratica_{nr}.pdf"
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", filetypes=[("PDF","*.pdf")],
            initialfile=default, title="Salva PDF")
        if not path: return
        try:
            genera_pdf(self._pratica_corrente, self._att_corrente or {}, path)
            messagebox.showinfo("PDF creato", f"File salvato:\n{path}")
            os.startfile(path)
        except Exception as e:
            messagebox.showerror("Errore PDF", str(e))

# ─────────────────────────────────────────────
#  APP PRINCIPALE
# ─────────────────────────────────────────────
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_NOME}  —  {OPERATORE}")
        self.geometry("1200x740")
        self.minsize(900, 580)
        self._excel_path      = ctk.StringVar(value=EXCEL_PATH_DEFAULT)
        self._appuntamento_var = ctk.StringVar(value="")  # persiste tra estrazioni
        self._server     = None
        self._banner     = None
        self._build()
        self._avvia_server()
        self._mostra_lista()
        self.after(800, self._controlla_scadenze)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _applica_tema(self):
        tema = SETTINGS.get("tema","light")
        ctk.set_appearance_mode("dark" if tema=="dark" else "light")

    def _build(self):
        self._applica_tema()
        T_ = T()

        top = ctk.CTkFrame(self, fg_color=T_["top_bg"], corner_radius=0, height=50)
        top.pack(fill="x"); top.pack_propagate(False)

        ctk.CTkLabel(top, text=f"  🏛  {APP_NOME}",
                     font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"),
                     text_color=T_["top_text"]).pack(side="left", padx=10)

        ctk.CTkEntry(top, textvariable=self._excel_path,
                     fg_color="#1a3f63", border_color=T_["blu_medio"],
                     text_color="white", height=30, width=220, corner_radius=6,
                     font=ctk.CTkFont(size=10)).pack(side="left", padx=(16,4))
        ctk.CTkButton(top, text="📁", width=32, height=30,
                      fg_color=T_["blu_medio"], hover_color="#1a3f63",
                      corner_radius=6, command=self._scegli_file).pack(side="left", padx=(0,8))

        ctk.CTkButton(top, text="📎  Bookmark",
                      command=lambda: PopupBookmarklet(self),
                      fg_color=T_["arancio"], hover_color="#C2410C",
                      height=30, corner_radius=8,
                      font=ctk.CTkFont(size=11)).pack(side="left", padx=(0,6))



        # Font size
        ctk.CTkButton(top, text="A-", command=self._font_down,
                      fg_color=T_["blu_medio"], hover_color=T_["blu_scuro"],
                      width=32, height=30, corner_radius=8,
                      font=ctk.CTkFont(size=11)).pack(side="left", padx=(0,2))
        ctk.CTkButton(top, text="A+", command=self._font_up,
                      fg_color=T_["blu_medio"], hover_color=T_["blu_scuro"],
                      width=32, height=30, corner_radius=8,
                      font=ctk.CTkFont(size=11)).pack(side="left", padx=(0,6))

        self.lbl_server = ctk.CTkLabel(top, text="⚪",
                                       font=ctk.CTkFont(size=10),
                                       text_color=T_["top_sub"])
        self.lbl_server.pack(side="left")

        ctk.CTkLabel(top, text=f"👤  {OPERATORE}",
                     font=ctk.CTkFont(size=11),
                     text_color=T_["top_sub"]).pack(side="right", padx=14)

        self._banner_frame = ctk.CTkFrame(self, fg_color="transparent", height=0)
        self._banner_frame.pack(fill="x")

        self._container = ctk.CTkFrame(self, fg_color=T_["sfondo"], corner_radius=0)
        self._container.pack(fill="both", expand=True)

        self._vista_lista  = VistaLista(self._container,
                                        get_path=self._excel_path.get,
                                        on_select=self._apri_pratica)
        self._vista_scheda = VistaScheda(self._container,
                                         get_path=self._excel_path.get,
                                         vai_lista=self._mostra_lista,
                                         appuntamento_var=self._appuntamento_var)

    def _font_up(self):
        SETTINGS["font_size"] = min(16, SETTINGS.get("font_size",12)+1)
        salva_settings(SETTINGS); self._ricostruisci()

    def _font_down(self):
        SETTINGS["font_size"] = max(9, SETTINGS.get("font_size",12)-1)
        salva_settings(SETTINGS); self._ricostruisci()

    def _ricostruisci(self):
        """Ricostruisce l'interfaccia con il nuovo tema/font."""
        self._applica_tema()
        nr_corrente  = self._vista_scheda._nr
        in_scheda    = self._vista_scheda.winfo_ismapped()
        data_app     = self._appuntamento_var.get()  # preserva data
        for w in self.winfo_children(): w.destroy()
        self._banner = None
        self._build()
        self._appuntamento_var.set(data_app)          # ripristina data
        self._avvia_server()
        if in_scheda and nr_corrente:
            self._apri_pratica(nr_corrente)
        else:
            self._mostra_lista()

    def _avvia_server(self):
        try:
            self._server = ServerLocale(callback=self._ricezione_dati)
            self._server.avvia()
            self.lbl_server.configure(text=f"🟢 :{SERVER_PORT}")
        except OSError:
            self.lbl_server.configure(text="🔴 Porta occupata")

    def _ricezione_dati(self, dati: dict):
        dati["Estratto Da"]    = OPERATORE
        dati["Data Estrazione"] = datetime.now().strftime("%d/%m/%Y %H:%M")
        self.after(0, self._gestisci_ricezione, dati)

    def _gestisci_ricezione(self, dati: dict):
        self.lift(); self.focus_force()
        # Aggiornamento automatico se siamo già sulla stessa pratica
        if self._vista_scheda.winfo_ismapped():
            if self._vista_scheda.aggiorna_dati(dati):
                return  # aggiornato silenziosamente
        # Altrimenti popup anteprima
        PopupAnteprima(self, dati, on_conferma=self._salva_e_apri)

    def _salva_e_apri(self, dati: dict):
        try:
            salva_pratica_excel(self._excel_path.get(), dati)
            nr = dati.get("Nr. Pratica","")
            self._mostra_lista()
            self.after(200, lambda: self._apri_pratica(nr))
        except Exception as e:
            messagebox.showerror("Errore", str(e))

    def _controlla_scadenze(self):
        path = self._excel_path.get()
        if not os.path.exists(path): return
        scad = pratiche_in_scadenza(path, 30)
        if scad and self._banner is None:
            self._banner_frame.configure(height=36)
            self._banner = BannerScadenze(
                self._banner_frame, scad,
                on_click=self._vai_scadenze)
            self._banner.pack(fill="x")

    def _vai_scadenze(self):
        self._mostra_lista()
        self.after(300, lambda: self._vista_lista._scad_var.set("30 giorni"))
        self.after(400, self._vista_lista._filtra)

    def _mostra_lista(self):
        self._vista_scheda.pack_forget()
        self._vista_lista.pack(fill="both", expand=True)
        self._vista_lista.carica()

    def _apri_pratica(self, nr: str):
        self._vista_lista.pack_forget()
        self._vista_scheda.pack(fill="both", expand=True)
        self._vista_scheda.carica(nr)

    def _scegli_file(self):
        p = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
        if p:
            self._excel_path.set(p)
            self._mostra_lista()

    def _on_close(self):
        if self._server: self._server.ferma()
        self.destroy()

# ─────────────────────────────────────────────
#  AVVIO CON SPLASH
# ─────────────────────────────────────────────
if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    splash = SplashScreen(root)
    splash.update()

    # Avvia l'app in background mentre splash è visibile
    def avvia():
        import time; time.sleep(2.5)
        root.after(0, _lancia_app)

    def _lancia_app():
        splash.chiudi()
        root.destroy()
        app = App()
        app.mainloop()

    threading.Thread(target=avvia, daemon=True).start()
    root.mainloop()
